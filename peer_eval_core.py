"""
Shared logic for Qualtrics peer evaluation Excel → graded workbook.

Used by `peer_eval_streamlit.py` (end users) and `peer_eval_notebook.ipynb` (development / batch runs).
"""
from __future__ import annotations

import io
import re
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RATING_MAP = {
    "Strongly agree": 5,
    "Somewhat agree": 4,
    "Neither agree nor disagree": 3,
    "Somewhat disagree": 2,
    "Strongly disagree": 1,
}

CRITERIA = [
    "Regularly attends meetings",
    "Actively participates and engages in team meetings",
    "Communicates well with others",
    "Contributes an equal amount or more time and energy to the projects as others",
    "Frequently goes above and beyond; takes initiative",
    "Maintains a can-do approach in frustrating situations",
    "Contributes meaningfully to group discussions",
    "I would not hesitate to work with this student again, given the opportunity",
]
COMMENT_LABEL = "Provide brief (UNIQUE) comments about this teammate."

HEADER_BG = "2E4057"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "EAF0FB"
TOTALS_BG = "D9E8FF"
SUMMARY_HDR = "1B3A5C"


def load_raw(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=None)
    codes = df.iloc[0].tolist()
    df.columns = codes
    df = df.iloc[2:].reset_index(drop=True)
    return df


def get_student_blocks(col_codes: list, col_labels: list):
    students = []
    seen = set()
    for code, label in zip(col_codes, col_labels):
        if not isinstance(label, str):
            continue

        m = re.match(r"^(\d+-\w+-\w+-[^-]+(?:-[^-]+)*) - REQUIRED - Provide brief", label)
        if not m:
            continue

        key = m.group(1)
        if key in seen:
            continue
        seen.add(key)

        comment_idx = col_codes.index(code)
        rating_codes = col_codes[comment_idx - 8 : comment_idx]

        parts = key.split("-")
        first = parts[1] if len(parts) > 1 else key

        students.append(
            {
                "key": key,
                "first": first,
                "rating_codes": rating_codes,
                "comment_code": code,
            }
        )

    return students


def build_dataframes(raw_df: pd.DataFrame, students: list) -> dict:
    name_col = "Q4"
    result = {}

    for s in students:
        rows = []
        for _, resp in raw_df.iterrows():
            reviewer = resp.get(name_col, "")

            scores = []
            for rc in s["rating_codes"]:
                raw_val = resp.get(rc, None)
                numeric = RATING_MAP.get(str(raw_val).strip(), None)
                scores.append(numeric)

            comment = resp.get(s["comment_code"], "")
            avg = (
                round(
                    sum(x for x in scores if x is not None)
                    / len([x for x in scores if x is not None]),
                    4,
                )
                if any(x is not None for x in scores)
                else None
            )

            row = [s["first"], reviewer] + scores + [avg, comment]
            rows.append(row)

        cols = ["Student", "Peer Reviewer"] + CRITERIA + ["Average Score", COMMENT_LABEL]
        df = pd.DataFrame(rows, columns=cols)
        df = df.sort_values("Peer Reviewer").reset_index(drop=True)
        result[s["key"]] = df

    return result


def compute_summary(student_dfs: dict, total_points: float) -> pd.DataFrame:
    rows = []
    for key, df in student_dfs.items():
        avg_scores = df["Average Score"].dropna()
        percent = avg_scores.mean() / 5.0
        total = round(percent * total_points, 6)
        rows.append({"Name": key, "Total Score": total})
    return pd.DataFrame(rows)


def _header_style(cell, bg=HEADER_BG, fg=HEADER_FG):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border():
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)


def write_student_sheet(ws, df: pd.DataFrame, total_points: float):
    columns = df.columns.tolist()
    ws.freeze_panes = "C2"

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _header_style(cell)
        cell.border = _border()

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == len(columns)))
            cell.border = _border()
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color=ALT_ROW_BG)

    avg_row = ws.max_row + 1
    criteria_start_col = 3
    avg_score_col = criteria_start_col + 8

    for col_idx in range(criteria_start_col, criteria_start_col + 9):
        col_letter = get_column_letter(col_idx)
        data_end = avg_row - 1
        formula = f"=AVERAGE({col_letter}2:{col_letter}{data_end})"
        cell = ws.cell(row=avg_row, column=col_idx, value=formula)
        cell.font = Font(bold=True, name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=TOTALS_BG)
        cell.number_format = "0.000"
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    avg_score_letter = get_column_letter(avg_score_col)
    percent_row = avg_row + 1
    ws.cell(row=percent_row, column=avg_score_col - 1, value="Percent Score").font = Font(
        bold=True, name="Arial", size=10
    )
    pct_cell = ws.cell(row=percent_row, column=avg_score_col, value=f"={avg_score_letter}{avg_row}/5")
    pct_cell.number_format = "0.0000"
    pct_cell.font = Font(bold=True, name="Arial", size=10)

    total_row = avg_row + 2
    ws.cell(row=total_row, column=avg_score_col - 1, value="Total Points").font = Font(
        bold=True, name="Arial", size=10
    )
    tot_cell = ws.cell(
        row=total_row,
        column=avg_score_col,
        value=f"={avg_score_letter}{percent_row}*{total_points}",
    )
    tot_cell.number_format = "0.00"
    tot_cell.font = Font(bold=True, name="Arial", size=10, color="1B3A5C")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for i in range(criteria_start_col, criteria_start_col + 9):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(len(columns))].width = 45

    ws.row_dimensions[1].height = 40


def write_summary_sheet(ws, summary_df: pd.DataFrame):
    for col_idx, col_name in enumerate(summary_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _header_style(cell, bg=SUMMARY_HDR)
        cell.border = _border()

    for row_idx, (_, row) in enumerate(summary_df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")
            cell.border = _border()
            if col_idx == 2:
                cell.number_format = "0.00"
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color=ALT_ROW_BG)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.row_dimensions[1].height = 28


def process_to_bytes(raw_path: str, total_points: float = 25.0) -> tuple[bytes, str]:
    """Run pipeline and return (xlsx bytes, suggested filename)."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)

        raw_df = load_raw(raw_path)
        df_headers = pd.read_excel(raw_path, header=None, nrows=2)
        col_codes = df_headers.iloc[0].tolist()
        col_labels = df_headers.iloc[1].tolist()

        students = get_student_blocks(col_codes, col_labels)
        if not students:
            raise ValueError(
                "No student blocks found. Export the raw Excel from Qualtrics "
                "(first row = codes, second = labels) and try again."
            )

        student_dfs = build_dataframes(raw_df, students)
        summary_df = compute_summary(student_dfs, total_points)

        base = Path(raw_path).stem
        suggested_name = f"{base}_OUTPUT.xlsx"

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        write_summary_sheet(ws_summary, summary_df)

        for s in students:
            key = s["key"]
            first = s["first"]
            df = student_dfs[key]
            ws = wb.create_sheet(title=first[:31])
            write_student_sheet(ws, df, total_points)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), suggested_name


def process(raw_path: str, total_points: float = 25.0, output_path: str | None = None) -> str:
    """Read Qualtrics Excel from disk and write `*_OUTPUT.xlsx` next to it (or to `output_path`)."""
    data, suggested = process_to_bytes(raw_path, total_points)
    out = output_path if output_path is not None else str(Path(raw_path).parent / suggested)
    Path(out).write_bytes(data)
    return out
